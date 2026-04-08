# -*- coding: utf-8 -*-
"""
测试数据管理类 - Data Layer
管理测试数据的读取和提供
"""
import json
import csv
import logging
from pathlib import Path
from typing import Dict, List, Any, Union
import openpyxl
from openpyxl import load_workbook

from config.settings import settings


class TestDataManager:
    """
    测试数据管理类
    支持JSON、Excel、CSV格式的测试数据
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.data_dir = Path(settings.DATA_DIR)
    
    def load_json(self, filename: str) -> Union[Dict, List]:
        """
        加载JSON文件
        :param filename: 文件名
        :return: JSON数据
        """
        file_path = self.data_dir / filename
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.logger.info(f"加载JSON文件成功: {file_path}")
            return data
        except Exception as e:
            self.logger.error(f"加载JSON文件失败: {file_path}, 错误: {e}")
            raise
    
    def save_json(self, filename: str, data: Union[Dict, List]) -> None:
        """
        保存数据到JSON文件
        :param filename: 文件名
        :param data: 要保存的数据
        """
        file_path = self.data_dir / filename
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.logger.info(f"保存JSON文件成功: {file_path}")
        except Exception as e:
            self.logger.error(f"保存JSON文件失败: {file_path}, 错误: {e}")
            raise
    
    def load_excel(self, filename: str, sheet_name: str = None) -> List[Dict]:
        """
        加载Excel文件
        :param filename: 文件名
        :param sheet_name: 工作表名称，默认第一个
        :return: 数据列表
        """
        file_path = self.data_dir / filename
        try:
            wb = load_workbook(file_path, data_only=True)
            
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            
            # 获取数据
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                # 过滤掉None值
                row_data = {k: v for k, v in row_data.items() if v is not None}
                data.append(row_data)
            
            self.logger.info(f"加载Excel文件成功: {file_path}, 共{len(data)}条数据")
            return data
        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {file_path}, 错误: {e}")
            raise
    
    def save_excel(self, filename: str, data: List[Dict], sheet_name: str = "Sheet1") -> None:
        """
        保存数据到Excel文件
        :param filename: 文件名
        :param data: 数据列表
        :param sheet_name: 工作表名称
        """
        file_path = self.data_dir / filename
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if data:
                # 写入表头
                headers = list(data[0].keys())
                ws.append(headers)
                
                # 写入数据
                for row_data in data:
                    row = [row_data.get(header, "") for header in headers]
                    ws.append(row)
            
            wb.save(file_path)
            self.logger.info(f"保存Excel文件成功: {file_path}")
        except Exception as e:
            self.logger.error(f"保存Excel文件失败: {file_path}, 错误: {e}")
            raise
    
    def load_csv(self, filename: str) -> List[Dict]:
        """
        加载CSV文件
        :param filename: 文件名
        :return: 数据列表
        """
        file_path = self.data_dir / filename
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                data = list(reader)
            self.logger.info(f"加载CSV文件成功: {file_path}, 共{len(data)}条数据")
            return data
        except Exception as e:
            self.logger.error(f"加载CSV文件失败: {file_path}, 错误: {e}")
            raise
    
    def get_login_data(self) -> List[Dict]:
        """
        获取登录测试数据
        :return: 登录测试数据列表
        """
        json_file = self.data_dir / "login_data.json"
        if json_file.exists():
            data = self.load_json("login_data.json")
            return data.get("login_cases", [])
        
        excel_file = self.data_dir / "login_data.xlsx"
        if excel_file.exists():
            return self.load_excel("login_data.xlsx")
        
        # 返回默认数据
        return [
            {"username": "admin", "password": "123456", "roleId": "1", "expected": "success"},
            {"username": "wrong_user", "password": "wrong_pass", "roleId": "1", "expected": "fail"},
        ]
    
    def get_activity_data(self) -> List[Dict]:
        """
        获取活动申请测试数据
        :return: 活动申请测试数据列表
        """
        json_file = self.data_dir / "activity_data.json"
        if json_file.exists():
            data = self.load_json("activity_data.json")
            return data.get("activity_cases", [])
        
        excel_file = self.data_dir / "activity_data.xlsx"
        if excel_file.exists():
            return self.load_excel("activity_data.xlsx")
        
        # 返回默认数据
        return [
            {
                "activity_place": "北京",
                "activity_type": "公司年会",
                "start_time": "2026-05-01",
                "end_time": "2026-05-02",
                "activity_people": "100",
                "activity_budget": "50000",
                "requirement_desc": "需要音响设备和舞台布置"
            }
        ]
    
    def get_employee_register_data(self) -> List[Dict]:
        """
        获取员工注册测试数据
        :return: 员工注册测试数据列表
        """
        json_file = self.data_dir / "employee_data.json"
        if json_file.exists():
            data = self.load_json("employee_data.json")
            return data.get("register_cases", [])
        
        excel_file = self.data_dir / "employee_data.xlsx"
        if excel_file.exists():
            return self.load_excel("employee_data.xlsx")
        
        # 返回默认数据
        return [
            {
                "username": "test_employee",
                "password": "123456",
                "employeeName": "测试员工",
                "phone": "13800138000",
                "sex": "男",
                "post": "仓库管理员",
                "city": "北京"
            }
        ]
    
    def get_material_data(self) -> List[Dict]:
        """
        获取物料管理测试数据
        :return: 物料管理测试数据列表
        """
        json_file = self.data_dir / "material_data.json"
        if json_file.exists():
            data = self.load_json("material_data.json")
            return data.get("material_cases", [])
        
        excel_file = self.data_dir / "material_data.xlsx"
        if excel_file.exists():
            return self.load_excel("material_data.xlsx")
        
        # 返回默认数据
        return [
            {
                "material_type": "专业音响",
                "quantity": "10",
                "price": "500",
                "category": "音响设备",
                "remarks": "高品质音响设备"
            }
        ]
    
    def get_reimbursement_data(self) -> List[Dict]:
        """
        获取报销申请测试数据
        :return: 报销申请测试数据列表
        """
        json_file = self.data_dir / "reimbursement_data.json"
        if json_file.exists():
            data = self.load_json("reimbursement_data.json")
            return data.get("reimbursement_cases", [])
        
        excel_file = self.data_dir / "reimbursement_data.xlsx"
        if excel_file.exists():
            return self.load_excel("reimbursement_data.xlsx")
        
        # 返回默认数据
        return [
            {
                "reimbursement_type": "1",
                "amount": "200",
                "expense_date": "2026-03-28",
                "description": "出差交通费"
            }
        ]


# 全局数据管理实例
data_manager = TestDataManager()
